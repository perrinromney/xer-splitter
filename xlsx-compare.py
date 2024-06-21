import openpyxl
from openpyxl import load_workbook

import subprocess
import tkinter as tk
from tkinter import filedialog
import platform


# Create a basic file selector GUI
root = tk.Tk()
root.withdraw()  # Hide the main window

# Ask the user to select the input .xer file
upload_file_path = filedialog.askopenfilename(title="Select .xlsx file of the upload", filetypes=[("XLSX Files", "*.xlsx")])
export_file_path = filedialog.askopenfilename(title="Select .xlsx file of the upload", filetypes=[("XLSX Files", "*.xlsx")])

upload_wb = load_workbook(upload_file_path)
export_wb = load_workbook(export_file_path)


